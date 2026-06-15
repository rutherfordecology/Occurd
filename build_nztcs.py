"""
Build nztcs.json from NZTCS Exported Data.xlsx

Produces a flat lookup keyed by lowercase scientific name.
Each entry uses the most-recently-assessed record (by Year Assessed).
Aliases (Previous Name, Alternative Names) also get index entries.

Output: nztcs.json (same directory as this script)
"""

import sys, json, re
sys.stdout.reconfigure(encoding='utf-8')
import openpyxl

XLSX = r'C:\Users\User\Downloads\NZTCS Exported Data.xlsx'
OUT  = r'C:\Users\User\Occurd\nztcs.json'

def clean(s):
    """Return stripped string or None."""
    if s is None:
        return None
    s = str(s).strip()
    return s if s else None

def norm_name(s):
    """
    Normalise a scientific name for lookup:
    - lowercase
    - strip leading/trailing quotes and whitespace
    - collapse internal whitespace
    """
    if not s:
        return None
    s = s.strip().strip('"').strip("'").lower()
    s = re.sub(r'\s+', ' ', s)
    # Strip authority — keep only genus + specific epithet (first 2 tokens)
    # BUT only if it looks like a binomial (2+ words). Uninomials stay as-is.
    tokens = s.split()
    if len(tokens) >= 2:
        # Heuristic: authority starts at first token that's not a lowercase word
        # Actually let's keep up to 3 tokens (subsp.) — GBIF names are usually 2
        # For robustness, just keep all tokens — we want to match exactly what GBIF returns
        pass
    return s

def split_alts(s):
    """Split Alternative Names field (comma or semicolon separated)."""
    if not s:
        return []
    parts = re.split(r'[;,]', str(s))
    return [p.strip() for p in parts if p.strip()]

print("Loading workbook…")
wb = openpyxl.load_workbook(XLSX, read_only=True, data_only=True)
ws = wb.active

headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
print(f"  Columns: {len(headers)}")

# Main lookup: key → entry dict
# We keep the most recent assessment per species name
lookup = {}   # lowercase name → entry
aliases = {}  # lowercase alias → canonical lowercase name

rows_read = 0

for row in ws.iter_rows(min_row=2, values_only=True):
    rows_read += 1
    d = dict(zip(headers, row))

    current_name = clean(d.get('Current Species Name'))
    if not current_name:
        continue

    year = d.get('Year Assessed')
    try:
        year = int(year) if year else None
    except (ValueError, TypeError):
        year = None

    entry = {
        'name':     current_name,
        'common':   clean(d.get('Preferred Common Name')),
        'maori':    clean(d.get('Preferred Māori Name')),
        'status':   clean(d.get('Status')),
        'category': clean(d.get('Category')),
        'year':     year,
        'bioStatus': clean(d.get('Bio Status')),
        'order':    clean(d.get('Order')),
        'family':   clean(d.get('Family')),
    }

    key = norm_name(current_name)
    if key:
        # Keep newer assessment if duplicate
        existing = lookup.get(key)
        if not existing or (year and (not existing['year'] or year > existing['year'])):
            lookup[key] = entry

        # Previous name → alias
        prev = clean(d.get('Previous Name'))
        if prev:
            pk = norm_name(prev)
            if pk and pk != key:
                aliases[pk] = key

        # Alternative names → aliases
        for alt in split_alts(d.get('Alternative Names')):
            ak = norm_name(alt)
            if ak and ak != key:
                aliases[ak] = key

wb.close()
print(f"  Rows read: {rows_read}")
print(f"  Unique names: {len(lookup)}")
print(f"  Aliases: {len(aliases)}")

# Merge aliases that don't shadow a real name
added = 0
for alias_key, canonical_key in aliases.items():
    if alias_key not in lookup and canonical_key in lookup:
        lookup[alias_key] = lookup[canonical_key]
        added += 1

print(f"  Alias entries added: {added}")

# ── Subspecies binomial + split-species + old-name aliases ────────────────
# Rule: only add a binomial alias when ALL subspecies that share that
# binomial have the SAME status.  Mixed-status subspecies (e.g. Tui with
# a Not Threatened mainland form and a NV Chatham Island form) must not
# generate an alias — showing the most-threatened one is misleading.

def is_latin_name(key, entry):
    """True if key uses only lowercase alpha/hyphen tokens and looks Latin."""
    parts = key.split()
    if len(parts) < 2:
        return False
    if not all(re.match(r'^[a-z][a-z\-]*$', p) for p in parts):
        return False
    return entry.get('name', '')[:1].isupper()

def has_latin_binomial(key, entry):
    """True if key starts with two pure-Latin lowercase tokens (genus + species),
    those tokens actually match the genus+species in the entry's name field,
    and the key has at least one more token after them.

    This covers:
      - standard subspecies: 'falco novaeseelandiae ferox'
      - informal/quoted varieties: 'falco novaeseelandiae "southern"'
      - authority-qualified synonyms: 'lycopodium fastigiatum r.br.'

    It rejects common-name keys like 'new zealand falcon' (tokens 'new'/'zealand'
    don't match the entry name 'Falco novaeseelandiae …').
    """
    parts = key.split()
    if len(parts) < 3:
        return False
    if not (re.match(r'^[a-z][a-z\-]*$', parts[0]) and re.match(r'^[a-z][a-z\-]*$', parts[1])):
        return False
    if not entry.get('name', '')[:1].isupper():
        return False
    # Verify the key's genus+species matches the entry's own name
    name_parts = norm_name(entry.get('name', '')).split()
    if len(name_parts) < 2:
        return False
    return name_parts[0] == parts[0] and name_parts[1] == parts[1]

def dedup_by_name(entries):
    """Deduplicate entry list by 'name' field, preserving order."""
    seen, out = set(), []
    for e in entries:
        n = e.get('name', '')
        if n not in seen:
            seen.add(n)
            out.append(e)
    return out

from collections import defaultdict

# --- 1. Subspecies / variety / informal-name binomial aliases ---
# Covers:
#   - Standard 3-token subspecies: "falco novaeseelandiae ferox"
#   - Informal/quoted varieties: 'falco novaeseelandiae "southern"', 'powelliphanta gilliesi "haidinger"'
#   - Author-qualified synonyms: "lycopodium fastigiatum r.br."
# Any key where the first two tokens are pure-Latin lowercase is a candidate for
# the 2-token binomial.  We no longer skip when the binomial already exists in
# lookup as an alias — that was suppressing conflict detection.
binomial_candidates = defaultdict(list)
for key, entry in lookup.items():
    if not has_latin_binomial(key, entry):
        continue
    parts = key.split()
    binomial = parts[0] + ' ' + parts[1]
    if key == binomial:
        continue  # don't add the binomial entry to its own candidate list
    binomial_candidates[binomial].append(entry)

subsp_added = subsp_skipped = 0
for binomial, raw_entries in binomial_candidates.items():
    entries = dedup_by_name(raw_entries)
    conflict_key = 'conflict:' + binomial
    statuses = set(e.get('status') for e in entries)
    if len(statuses) == 1:
        # All subspecies agree — add alias only if nothing exists yet
        if binomial not in lookup:
            lookup[binomial] = entries[0]
            subsp_added += 1
    else:
        # Mixed statuses — write conflict regardless of whether an alias exists
        subsp_skipped += 1
        if conflict_key not in lookup:
            lookup[conflict_key] = {'_conflict': True, 'options': entries}

print(f"  Subspecies/variety binomial aliases added: {subsp_added} (skipped {subsp_skipped} mixed-status)")

# --- 2. Old-name binomial aliases (4+ token alias keys → 2-word binomial) ---
# e.g. "porzana tabuensis tabuensis gmelin, 1789" → "porzana tabuensis"
# Section 1 now handles most of these via the widened has_latin_binomial check,
# but keep section 2 for stragglers that only appear after the alias expansion pass.
old_binomial_candidates = defaultdict(list)
for key in list(lookup.keys()):
    entry = lookup[key]
    if not has_latin_binomial(key, entry):
        continue
    parts = key.split()
    binomial = parts[0] + ' ' + parts[1]
    if key == binomial or binomial in lookup or 'conflict:' + binomial in lookup:
        continue
    old_binomial_candidates[binomial].append(entry)

old_binomial_added = old_binomial_skipped = 0
for binomial, raw_entries in old_binomial_candidates.items():
    entries = dedup_by_name(raw_entries)
    statuses = set(e.get('status') for e in entries)
    if len(statuses) == 1:
        lookup[binomial] = entries[0]
        old_binomial_added += 1
    else:
        old_binomial_skipped += 1
        ck = 'conflict:' + binomial
        if ck not in lookup:
            lookup[ck] = {'_conflict': True, 'options': entries}

print(f"  Old-name binomial aliases added: {old_binomial_added} (skipped {old_binomial_skipped} mixed-status)")

# --- 3. Split-species aliases (genus + subspecies epithet, e.g. "zapornia affinis") ---
# When a subspecies is treated as a full species by some authorities.
# Only safe when the subspecies epithet maps to a unique status.
split_candidates = defaultdict(list)
for key in list(lookup.keys()):
    parts = key.split()
    if len(parts) != 3:
        continue
    if not is_latin_name(key, lookup[key]):
        continue
    genus, species_ep, subsp_ep = parts
    if subsp_ep == species_ep:
        continue  # nominate subspecies already covered by binomial alias
    split_key = genus + ' ' + subsp_ep
    if split_key not in lookup:
        split_candidates[split_key].append(lookup[key])

split_added = split_skipped = 0
for split_key, entries in split_candidates.items():
    statuses = set(e.get('status') for e in entries)
    if len(statuses) == 1:
        lookup[split_key] = entries[0]
        split_added += 1
    else:
        split_skipped += 1
        ck = 'conflict:' + split_key
        if ck not in lookup:
            lookup[ck] = {'_conflict': True, 'options': entries}

print(f"  Split-species aliases added: {split_added} (skipped {split_skipped} mixed-status)")
print(f"  Total lookup entries: {len(lookup)}")

# Status summary
from collections import Counter
status_counts = Counter(v['status'] for v in lookup.values() if 'status' in v)
print("\nStatus distribution (unique entries):")
for status, count in status_counts.most_common():
    print(f"  {status}: {count}")

# Write JSON
print(f"\nWriting {OUT}…")
with open(OUT, 'w', encoding='utf-8') as f:
    json.dump(lookup, f, ensure_ascii=False, separators=(',', ':'))

import os
size_kb = os.path.getsize(OUT) / 1024
print(f"Done. File size: {size_kb:.1f} KB")
